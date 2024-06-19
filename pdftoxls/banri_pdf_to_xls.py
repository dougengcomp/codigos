import PyPDF2
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with open(pdf_path, "rb") as file:
            reader = PyPDF2.PdfReader(file)
            num_pages = len(reader.pages)
            
            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text += page.extract_text()
                
    except Exception as e:
        print(f"Error reading PDF file: {e}")
        
    return text

def parse_payments_section(text, section_title):
    section_regex = re.compile(rf"{section_title}.*?(?=(?:PARCELAS|Extrato|$))", re.DOTALL)
    section_text = section_regex.search(text)
    if not section_text:
        return []

    pattern = re.compile(r"(\d{2}/\d{2}/\d{4})\s+[^0-9]+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)")
    matches = pattern.findall(section_text.group(0))
    return matches

def save_to_excel(workbook, sheet_name, data, headers):
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(headers)

    amount_style_name = "amount_style"
    amount_style = NamedStyle(name=amount_style_name, number_format='#,##0.00')

    try:
        workbook.add_named_style(amount_style)
    except ValueError:
        # If style already exists, retrieve it
        amount_style = workbook._named_styles[amount_style_name]

    for row in data:
        parsed_row = [row[0]] + [float(item.replace('.', '').replace(',', '.')) for item in row[1:]]
        sheet.append(parsed_row)
        for col in range(2, len(parsed_row) + 1):
            sheet.cell(row=sheet.max_row, column=col).style = amount_style

def main(pdf_filename):
    pdf_path = Path(__file__).parent / pdf_filename
    text = extract_text_from_pdf(pdf_path)

    if text:
        workbook = Workbook()
        workbook.remove(workbook.active)

        # Extract and save "PARCELAS PAGAS" section
        pagas_data = parse_payments_section(text, "PARCELAS PAGAS")
        if pagas_data:
            save_to_excel(workbook, "Parcelas Pagas", pagas_data, ['Date', 'Principal (R$)', 'Juros/Encargos (R$)', 'Total (R$)'])

        # Extract and save "PARCELAS A VENCER" section
        a_vencer_data = parse_payments_section(text, "PARCELAS A VENCER")
        if a_vencer_data:
            save_to_excel(workbook, "Parcelas a Vencer", a_vencer_data, ['Date', 'Principal (R$)', 'Juros/Encargos (R$)', 'Total (R$)'])

        excel_file_path = Path(__file__).parent / Path(pdf_filename).stem
        excel_file_path = excel_file_path.with_suffix('.xlsx')
        workbook.save(excel_file_path)
        print(f"Data saved to: {excel_file_path}")
        return excel_file_path
    else:
        print("No text extracted from the PDF.")
        return None

# PDF file name (assuming it's in the same directory as the script)
pdf_filename = "ESPELHO ADIR 6.pdf"

# Execute the main function
output_file_path = main(pdf_filename)
print(output_file_path)
