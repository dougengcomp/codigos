import PyPDF2
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

# Define the amount style
amount_style = NamedStyle(name="amount_style", number_format='#,##0.00')

def extract_text_from_pdf(pdf_path):
    text = ""
    try:
        with open(pdf_path, "rb") as file:
            reader = PyPDF2.PdfReader(file)
            num_pages = len(reader.pages)
            
            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text += page.extract_text()
                
    except Exception as e:
        print(f"Error reading PDF file: {e}")
        
    return text

def search_keyword_date_and_amount_in_text(text):
    pattern = r"Mês de referência:\s*(\d{2}/\d{4})\s*R\$\s*([\d.,]+)"
    matches = re.findall(pattern, text)
    return matches

def search_banco_occurrences(text):
    pattern = r"\b(\d{2}/\d{4})\b.*?BANCO DO ESTADO DO RIO\s*GRANDE DO SUL S.A.\s*R\$\s*([\d.,]+)"
    banco_matches = re.findall(pattern, text, re.DOTALL)
    return banco_matches

def save_to_excel_total(workbook, matches):
    sheet = workbook.create_sheet(title='Dividas')
    sheet.append(['Date', 'Total Amount'])

    for date, amount in matches:
        amount = amount.replace('.', '').replace(',', '.')  # Replacing '.' with '' and ',' with '.'
        try:
            amount = float(amount)
            sheet.append([date, amount])
            sheet.cell(row=sheet.max_row, column=2).style = amount_style
        except ValueError:
            print(f"Error converting amount to float: {amount}")

def save_to_excel_banri(workbook, banco_matches):
    sheet = workbook.create_sheet(title='Banri')
    sheet.append(['Date', 'Total Amount Banri'])

    for date, amount in banco_matches:
        amount = amount.replace('.', '').replace(',', '.')  # Replacing '.' with '' and ',' with '.'
        try:
            amount = float(amount)
            sheet.append([date, amount])
            sheet.cell(row=sheet.max_row, column=2).style = amount_style
        except ValueError:
            print(f"Error converting amount to float: {amount}")

def main(pdf_path):
    script_dir = Path(__file__).resolve().parent
    text = extract_text_from_pdf(pdf_path)
    
    if text:
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet created with the workbook

        # Add the named style to the workbook
        workbook.add_named_style(amount_style)

        matches = search_keyword_date_and_amount_in_text(text)
        if matches:
            print(f"Found {len(matches)} occurrence(s) of the keyword 'Mês de referência:'.")
            save_to_excel_total(workbook, matches)
        else:
            print("No occurrences of the keyword 'Mês de referência:' found.")
        
        banco_matches = search_banco_occurrences(text)
        if banco_matches:            
            print(f"Found {len(banco_matches)} occurrence(s) of the keyword 'Banri'.")
            save_to_excel_banri(workbook, banco_matches)
        else:
            print("No occurrences of the keyword 'Banri' found.")

        excel_file_path = script_dir / 'dividas_combined.xlsx'
        workbook.save(excel_file_path)
        print(f"Data saved to: {excel_file_path}")
    else:
        print("No text extracted from the PDF.")

# Path to the PDF file
pdf_path = Path(__file__).with_name("dividas.pdf")

if __name__ == "__main__":
    main(pdf_path)
